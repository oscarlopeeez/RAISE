import openpyxl
from django.http import HttpResponse
from datetime import datetime

# Mapeo de códigos de escenarios
SCENARIOS = {
    "eve_base": {"codigo": "0030", "columna": 6},
    "eve_parallel_up": {"codigo": "0040", "columna": 7},
    "eve_parallel_down": {"codigo": "0050", "columna": 8},
    "eve_steepener": {"codigo": "0060", "columna": 9},
    "eve_flattener": {"codigo": "0070", "columna": 10},
    "eve_short_up": {"codigo": "0080", "columna": 11},
    "eve_short_down": {"codigo": "0090", "columna": 12},
    "nii_base": {"codigo": "0120", "columna": 13},
    "nii_parallel_up": {"codigo": "0130", "columna": 14},
    "nii_parallel_down": {"codigo": "0140", "columna": 15},
}

# Mapeo de productos activos
PRODUCTS_ACTIVOS = {
    "Central bank": {"codigo": "0030", "fila": 13},
    "Interbank": {"codigo": "0040", "fila": 14},
    "Loans and advances": {"codigo": "0050", "fila": 15},
    "Debt securities": {"codigo": "0120", "fila": 16},
    "Derivatives hedging assets": {"codigo": "0140", "fila": 17},
    "Hedging debt securities": {"codigo": "0160", "fila": 18},
    "Hedging other assets": {"codigo": "0170", "fila": 19},
    "Other": {"codigo": "0180", "fila": 20},
}

# Mapeo de productos pasivos
PRODUCTS_PASIVOS = {
    "Central bank": {"codigo": "0220", "fila": 23},
    "Interbank": {"codigo": "0230", "fila": 24},
    "Debt securities issued": {"codigo": "0240", "fila": 25},
    "NMDs: Retail transactional": {"codigo": "0270", "fila": 26},
    "NMDs: Retail non-transactional": {"codigo": "0310", "fila": 27},
    "NMDs: Wholesale non-financial": {"codigo": "0350", "fila": 28},
    "NMDs: Wholesale financial": {"codigo": "0390", "fila": 29},
    "Term deposits": {"codigo": "0420", "fila": 30},
    "Derivatives hedging liabilities": {"codigo": "0470", "fila": 31},
    "Hedging debt securities": {"codigo": "0490", "fila": 32},
    "Hedging other liabilities": {"codigo": "0500", "fila": 33},
    "Other": {"codigo": "0510", "fila": 34},
}

def export_excel(fecha, activos, pasivos, banco_name):
    wb = openpyxl.load_workbook(filename='template.xlsx')
    ws = wb.active

    # Rellenar info básica
    ws["C3"] = banco_name
    ws["C4"] = "EUR"
    ws["C5"] = fecha
    
    for producto, data in activos.items():
        if producto in PRODUCTS_ACTIVOS:
            fila = PRODUCTS_ACTIVOS[producto]["fila"]
            scenario = data.get("scenario", {})
            for field_name, scenario_info in SCENARIOS.items():
                columna = scenario_info["columna"]
                valor = scenario.get(field_name, 0)
                ws.cell(row=fila, column=columna).value = abs(valor)
    
    for producto, data in pasivos.items():
        if producto in PRODUCTS_PASIVOS:
            fila = PRODUCTS_PASIVOS[producto]["fila"]
            scenario = data.get("scenario", {})
            for field_name, scenario_info in SCENARIOS.items():
                columna = scenario_info["columna"]
                valor = scenario.get(field_name, 0)
                ws.cell(row=fila, column=columna).value = abs(valor)
    
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'J_03_00_{banco_name}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)

    return response


